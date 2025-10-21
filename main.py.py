import sys
import os
from PyQt5 import QtWidgets, uic
from openpyxl import Workbook, load_workbook
from datetime import datetime

project_path = r"C:\Users\TUF\Desktop\Python-Excel-DataEntry"
os.chdir(project_path)

class IndustryLogger(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi("form_ui.ui", self)  # Load your Qt Designer UI file

        # Connect the button click event
        self.button_save.clicked.connect(self.save_data)

    def save_data(self):
        try:
            # Get input values from the GUI
            machine_id = self.input_machine.text()
            operator = self.input_operator.text()
            status = self.status.currentText()
            output = self.input_output.text()

            # Validation checks
            if not machine_id or not operator or not output:
                QtWidgets.QMessageBox.warning(self, "Warning", "Please fill in all fields!")
                return

            try:
                output = int(output)
            except ValueError:
                QtWidgets.QMessageBox.warning(self, "Error", "Output must be a number!")
                return

            # Create or open Excel file
            file_path = "production_log.xlsx"
            if not os.path.exists(file_path):
                wb = Workbook()
                ws = wb.active
                ws.title = "Production Log"
                ws.append(["Timestamp", "Machine ID", "Operator", "Status", "Output (pcs)"])
                wb.save(file_path)

            # Load workbook and append data
            wb = load_workbook(file_path)
            ws = wb.active

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([timestamp, machine_id, operator, status, output])
            wb.save(file_path)

            # Show success message
            QtWidgets.QMessageBox.information(self, "Success", "Data saved successfully!")

            # Clear inputs for next entry
            self.input_machine.clear()
            self.input_operator.clear()
            self.input_output.clear()
            self.status.setCurrentIndex(0)

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", str(e))

# Run the App
app = QtWidgets.QApplication(sys.argv)
window = IndustryLogger()
window.show()
sys.exit(app.exec_())